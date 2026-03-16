from __future__ import annotations

import html
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT_DIR = Path(__file__).resolve().parent.parent
REPORTS_DIR = ROOT_DIR / "reports"
INPUT_FILES = {
    "backend": REPORTS_DIR / "backend-junit.xml",
    "frontend": REPORTS_DIR / "frontend-junit.xml",
}
METADATA_FILES = {
    "backend": ROOT_DIR / "tests" / "metadata" / "backend-test-catalog.json",
    "frontend": ROOT_DIR / "frontend" / "test-metadata.json",
}
OUTPUT_FILE = REPORTS_DIR / "test-results.xlsx"
HTML_OUTPUT_FILE = REPORTS_DIR / "test-dashboard.html"
STATUS_FILLS = {
    "passed": PatternFill(fill_type="solid", fgColor="E6F4EA"),
    "failed": PatternFill(fill_type="solid", fgColor="FCE8E6"),
    "skipped": PatternFill(fill_type="solid", fgColor="FFF4D6"),
}
BONUS_PREFIX_PATTERN = re.compile(r"^\s*Bonus:\s*", re.IGNORECASE)


def normalize_coverage_label(label: Any) -> str:
    normalized = BONUS_PREFIX_PATTERN.sub("", str(label).strip())
    return normalized or "Uncategorized"


def load_metadata(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def parse_junit(path: Path, suite_name: str, metadata: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    root = ElementTree.parse(path).getroot()
    rows: list[dict[str, Any]] = []

    for testcase in root.iter("testcase"):
        failure = testcase.find("failure")
        error = testcase.find("error")
        skipped = testcase.find("skipped")

        status = "passed"
        details = ""
        if failure is not None:
            status = "failed"
            details = " ".join(
                filter(
                    None,
                    [
                        failure.attrib.get("message", "").strip(),
                        (failure.text or "").strip(),
                    ],
                )
            )
        elif error is not None:
            status = "failed"
            details = " ".join(
                filter(
                    None,
                    [
                        error.attrib.get("message", "").strip(),
                        (error.text or "").strip(),
                    ],
                )
            )
        elif skipped is not None:
            status = "skipped"
            details = skipped.attrib.get("message", "").strip() or (skipped.text or "").strip()

        classname = testcase.attrib.get("classname", "")
        name = testcase.attrib.get("name", "")
        key = f"{classname}::{name}"
        meta = metadata.get(key, {})
        description = meta.get("description") or name.replace("_", " ")
        raw_coverage = meta.get("coverage", [])
        if not isinstance(raw_coverage, list):
            raw_coverage = [raw_coverage]
        coverage = [normalize_coverage_label(area) for area in raw_coverage]

        rows.append(
            {
                "suite": suite_name,
                "key": key,
                "classname": classname,
                "name": name,
                "description": description,
                "coverage": coverage,
                "status": status,
                "time": testcase.attrib.get("time", "0"),
                "details": details,
            }
        )

    return rows


def collect_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for suite_name, report_path in INPUT_FILES.items():
        suite_rows = parse_junit(report_path, suite_name, load_metadata(METADATA_FILES[suite_name]))
        rows.extend(suite_rows)
    return rows


def build_workbook(rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["suite", "tests", "passed", "failed", "skipped"])

    grouped_by_suite: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped_by_suite[row["suite"]].append(row)

    for suite_name, suite_rows in grouped_by_suite.items():
        status_counter = Counter(row["status"] for row in suite_rows)
        summary_sheet.append(
            [
                suite_name,
                len(suite_rows),
                status_counter.get("passed", 0),
                status_counter.get("failed", 0),
                status_counter.get("skipped", 0),
            ]
        )

    coverage_sheet = workbook.create_sheet(title="coverage")
    coverage_sheet.append(["coverage_area", "tests"])
    coverage_counter = Counter(area for row in rows for area in row["coverage"])
    for area, count in coverage_counter.most_common():
        coverage_sheet.append([area, count])

    for suite_name, suite_rows in grouped_by_suite.items():
        suite_sheet = workbook.create_sheet(title=suite_name)
        suite_sheet.append(["key", "classname", "name", "description", "coverage", "status", "time", "details"])
        for row in suite_rows:
            suite_sheet.append(
                [
                    row["key"],
                    row["classname"],
                    row["name"],
                    row["description"],
                    ", ".join(row["coverage"]),
                    row["status"],
                    row["time"],
                    row["details"],
                ]
            )
            suite_sheet.cell(row=suite_sheet.max_row, column=6).fill = STATUS_FILLS.get(row["status"], PatternFill())
        for cell in suite_sheet[1]:
            cell.font = Font(bold=True)

    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    for cell in coverage_sheet[1]:
        cell.font = Font(bold=True)

    return workbook


def build_dashboard(rows: list[dict[str, Any]]) -> str:
    total = len(rows)
    status_counter = Counter(row["status"] for row in rows)
    coverage_counter = Counter(area for row in rows for area in row["coverage"])
    suite_counter: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows:
        suite_counter[row["suite"]][row["status"]] += 1

    summary_cards = "\n".join(
        [
            _summary_card("Total tests", total, "Complete backend and frontend validation set"),
            _summary_card("Passed", status_counter.get("passed", 0), "Stable validations currently passing"),
            _summary_card("Failed", status_counter.get("failed", 0), "Tests that need attention"),
            _summary_card("Skipped", status_counter.get("skipped", 0), "Cases intentionally or conditionally skipped"),
        ]
    )

    suite_blocks = "\n".join(
        f"""
        <article class="metric-card metric-card--compact">
          <div class="metric-card__label">{html.escape(suite)}</div>
          <strong>{sum(counter.values())}</strong>
          <p>{counter.get('passed', 0)} passed &bull; {counter.get('failed', 0)} failed &bull; {counter.get('skipped', 0)} skipped</p>
        </article>
        """
        for suite, counter in suite_counter.items()
    )

    coverage_blocks = "\n".join(
        f'<li><span>{html.escape(area)}</span><strong>{count}</strong></li>'
        for area, count in coverage_counter.most_common()
    )

    table_rows = "\n".join(
        f"""
        <tr data-search="{html.escape(' '.join([row['suite'], row['classname'], row['name'], row['description'], ' '.join(row['coverage'])]).lower())}">
          <td>{html.escape(row['suite'])}</td>
          <td>{html.escape(row['classname'])}</td>
          <td>{html.escape(row['name'])}</td>
          <td>{html.escape(row['description'])}</td>
          <td>{html.escape(', '.join(row['coverage']))}</td>
          <td><span class="status-pill status-pill--{row['status']}">{html.escape(row['status'])}</span></td>
          <td>{html.escape(str(row['time']))}</td>
          <td>{html.escape(row['details'] or '-')}</td>
        </tr>
        """
        for row in rows
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>TaskBuddy Test Dashboard</title>
  <style>
    :root {{
      color-scheme: light;
      font-family: "Segoe UI Variable Text", "Segoe UI", Aptos, sans-serif;
      --bg: #f3f7fc;
      --bg-strong: #e8f0fb;
      --surface: rgba(255, 255, 255, 0.97);
      --surface-strong: #ffffff;
      --line: rgba(15, 58, 95, 0.12);
      --text: #14385b;
      --muted: #6782a0;
      --navy: #0e416c;
      --blue: #0f6eb4;
      --green: #207a44;
      --red: #c44d43;
      --yellow: #a97816;
      --shadow: 0 18px 40px rgba(13, 47, 82, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(157, 195, 234, 0.18), transparent 28%),
        linear-gradient(180deg, #fbfdff 0%, var(--bg) 50%, #edf4fb 100%);
    }}
    .shell {{
      max-width: 1440px;
      margin: 0 auto;
      padding: 28px 20px 36px;
    }}
    .hero {{
      display: grid;
      gap: 18px;
      margin-bottom: 24px;
    }}
    .hero__card,
    .panel,
    .results {{
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 28px;
      box-shadow: var(--shadow);
      backdrop-filter: blur(10px);
    }}
    .hero__card {{
      padding: 24px;
      background:
        radial-gradient(circle at top right, rgba(177, 207, 242, 0.26), transparent 20%),
        linear-gradient(145deg, rgba(255, 255, 255, 0.98), rgba(247, 251, 255, 0.98));
    }}
    .hero__eyebrow {{
      margin: 0 0 6px;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      font-size: 0.82rem;
      color: var(--muted);
    }}
    h1, h2 {{ margin: 0; color: var(--navy); }}
    .hero__copy {{
      margin: 10px 0 0;
      max-width: 56ch;
      color: var(--muted);
      line-height: 1.6;
    }}
    .summary-grid,
    .suite-grid {{
      display: grid;
      gap: 14px;
    }}
    .summary-grid {{
      grid-template-columns: repeat(4, minmax(0, 1fr));
    }}
    .suite-grid {{
      grid-template-columns: repeat(2, minmax(0, 1fr));
    }}
    .metric-card {{
      padding: 18px;
      border-radius: 22px;
      border: 1px solid var(--line);
      background: rgba(255, 255, 255, 0.96);
    }}
    .metric-card strong {{
      display: block;
      margin-top: 10px;
      font-size: clamp(1.6rem, 2vw, 2rem);
      color: var(--navy);
    }}
    .metric-card p,
    .metric-card__label,
    .panel p,
    .toolbar__meta,
    td {{
      color: var(--muted);
    }}
    .metric-card p {{
      margin: 8px 0 0;
      line-height: 1.5;
    }}
    .metric-card--compact strong {{
      margin-top: 8px;
      font-size: 1.5rem;
    }}
    .content-grid {{
      display: grid;
      grid-template-columns: minmax(0, 1.15fr) minmax(280px, 0.85fr);
      gap: 16px;
      margin-bottom: 20px;
    }}
    .panel {{
      padding: 18px;
    }}
    .panel__header {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 12px;
    }}
    .panel ul {{
      list-style: none;
      margin: 0;
      padding: 0;
      display: grid;
      gap: 10px;
    }}
    .panel li {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      padding: 12px 14px;
      border-radius: 16px;
      border: 1px solid var(--line);
      background: linear-gradient(180deg, rgba(249, 252, 255, 0.98), rgba(245, 249, 255, 0.98));
    }}
    .panel li span {{
      color: var(--text);
      font-weight: 600;
    }}
    .results {{
      overflow: hidden;
    }}
    .toolbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      padding: 20px;
      border-bottom: 1px solid var(--line);
      background: linear-gradient(180deg, rgba(249, 252, 255, 0.98), rgba(243, 248, 255, 0.98));
    }}
    .toolbar__meta {{
      margin-top: 6px;
      font-size: 0.9rem;
    }}
    input[type="search"] {{
      width: min(420px, 100%);
      border: 1px solid var(--line);
      border-radius: 16px;
      padding: 13px 14px;
      background: var(--surface-strong);
      font: inherit;
      color: var(--text);
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      background: var(--surface-strong);
    }}
    th, td {{
      text-align: left;
      padding: 14px 16px;
      border-bottom: 1px solid rgba(15, 58, 95, 0.08);
      vertical-align: top;
      font-size: 0.92rem;
    }}
    th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: rgba(237, 245, 255, 0.97);
      color: var(--muted);
      font-size: 0.76rem;
      letter-spacing: 0.08em;
      text-transform: uppercase;
    }}
    tbody tr:hover {{
      background: rgba(244, 249, 255, 0.72);
    }}
    .status-pill {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 78px;
      padding: 6px 10px;
      border-radius: 999px;
      font-weight: 700;
      text-transform: capitalize;
      border: 1px solid transparent;
    }}
    .status-pill--passed {{
      color: var(--green);
      background: rgba(223, 240, 221, 0.9);
      border-color: rgba(32, 122, 68, 0.12);
    }}
    .status-pill--failed {{
      color: var(--red);
      background: rgba(255, 239, 236, 0.92);
      border-color: rgba(196, 77, 67, 0.14);
    }}
    .status-pill--skipped {{
      color: var(--yellow);
      background: rgba(255, 244, 214, 0.92);
      border-color: rgba(169, 120, 22, 0.12);
    }}
    @media (max-width: 1080px) {{
      .summary-grid,
      .suite-grid,
      .content-grid {{
        grid-template-columns: 1fr;
      }}
      .toolbar {{
        flex-direction: column;
        align-items: stretch;
      }}
    }}
    @media (max-width: 760px) {{
      .shell {{
        padding: 18px 12px 24px;
      }}
      .hero__card,
      .panel,
      .results {{
        border-radius: 22px;
      }}
      th, td {{
        padding: 12px;
      }}
    }}
  </style>
</head>
<body>
  <div class="shell">
    <section class="hero">
      <article class="hero__card">
        <p class="hero__eyebrow">TaskBuddy review pack</p>
        <h1>Automated Test Dashboard</h1>
        <p class="hero__copy">Backend and frontend verification results for the TaskBuddy coding challenge submission, including normalized coverage mapping, per-test descriptions, and searchable failure detail.</p>
      </article>
      <div class="summary-grid">
        {summary_cards}
      </div>
    </section>

    <section class="content-grid">
      <article class="panel">
        <div class="panel__header">
          <div>
            <h2>Suite Breakdown</h2>
            <p>High-level view of backend and frontend validation results.</p>
          </div>
        </div>
        <div class="suite-grid">
          {suite_blocks}
        </div>
      </article>

      <article class="panel">
        <div class="panel__header">
          <div>
            <h2>Coverage Areas</h2>
            <p>Normalized challenge coverage labels used in the export outputs.</p>
          </div>
        </div>
        <ul>{coverage_blocks}</ul>
      </article>
    </section>

    <section class="results">
      <div class="toolbar">
        <div>
          <h2>Detailed Results</h2>
          <p class="toolbar__meta">Search by suite, class, test name, description, or coverage area.</p>
        </div>
        <input id="search" type="search" placeholder="Search tests, descriptions, or coverage..." />
      </div>

      <table>
        <thead>
          <tr>
            <th>Suite</th>
            <th>Class</th>
            <th>Test</th>
            <th>Description</th>
            <th>Coverage</th>
            <th>Status</th>
            <th>Time</th>
            <th>Failure / Notes</th>
          </tr>
        </thead>
        <tbody id="results-table">
          {table_rows}
        </tbody>
      </table>
    </section>
  </div>
  <script>
    const searchBox = document.getElementById('search');
    const rows = Array.from(document.querySelectorAll('#results-table tr'));
    searchBox?.addEventListener('input', (event) => {{
      const query = String(event.target.value || '').toLowerCase().trim();
      for (const row of rows) {{
        const haystack = row.getAttribute('data-search') || '';
        row.style.display = !query || haystack.includes(query) ? '' : 'none';
      }}
    }});
  </script>
</body>
</html>"""


def _summary_card(title: str, value: int, description: str) -> str:
    return f"""
    <article class="metric-card">
      <div class="metric-card__label">{html.escape(title)}</div>
      <strong>{value}</strong>
      <p>{html.escape(description)}</p>
    </article>
    """


def main() -> None:
    rows = collect_rows()
    workbook = build_workbook(rows)
    workbook.save(OUTPUT_FILE)
    HTML_OUTPUT_FILE.write_text(build_dashboard(rows), encoding="utf-8")


if __name__ == "__main__":
    main()
